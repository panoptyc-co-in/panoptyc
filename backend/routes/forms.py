from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from datetime import datetime
from openpyxl import Workbook, load_workbook
from pathlib import Path
import os

router = APIRouter()

# Ensure data directory exists
DATA_DIR = Path("/app/backend/data")
DATA_DIR.mkdir(exist_ok=True)

APPLICATIONS_FILE = DATA_DIR / "applications.xlsx"
PROFILE_SETUP_FILE = DATA_DIR / "profile_setup.xlsx"


class ApplicationForm(BaseModel):
    fullName: str
    phone: str
    city: str


class ProfileSetupForm(BaseModel):
    email: str
    password: str


def init_applications_excel():
    """Initialize applications Excel file with headers if it doesn't exist"""
    if not APPLICATIONS_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        ws.append(["Timestamp", "Full Name", "Phone Number", "City"])
        wb.save(APPLICATIONS_FILE)


def init_profile_setup_excel():
    """Initialize profile setup Excel file with headers if it doesn't exist"""
    if not PROFILE_SETUP_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Profile Setup"
        ws.append(["Timestamp", "Email", "Password"])
        wb.save(PROFILE_SETUP_FILE)


@router.post("/api/submit-application")
async def submit_application(form: ApplicationForm):
    """Save application form data to Excel"""
    try:
        # Initialize Excel file if it doesn't exist
        init_applications_excel()
        
        # Load workbook and get active sheet
        wb = load_workbook(APPLICATIONS_FILE)
        ws = wb.active
        
        # Append new row with timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([timestamp, form.fullName, form.phone, form.city])
        
        # Save workbook
        wb.save(APPLICATIONS_FILE)
        
        return {
            "success": True,
            "message": "Application submitted successfully!",
            "data": {
                "fullName": form.fullName,
                "timestamp": timestamp
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save application: {str(e)}")


@router.post("/api/submit-profile-setup")
async def submit_profile_setup(form: ProfileSetupForm):
    """Save profile setup form data to Excel"""
    try:
        # Initialize Excel file if it doesn't exist
        init_profile_setup_excel()
        
        # Load workbook and get active sheet
        wb = load_workbook(PROFILE_SETUP_FILE)
        ws = wb.active
        
        # Append new row with timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([timestamp, form.email, form.password])
        
        # Save workbook
        wb.save(PROFILE_SETUP_FILE)
        
        return {
            "success": True,
            "message": "Profile setup submitted successfully!",
            "data": {
                "email": form.email,
                "timestamp": timestamp
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save profile setup: {str(e)}")


@router.get("/api/download-applications")
async def download_applications():
    """Get path to applications Excel file"""
    if not APPLICATIONS_FILE.exists():
        raise HTTPException(status_code=404, detail="No applications found")
    
    return {
        "success": True,
        "file_path": str(APPLICATIONS_FILE),
        "message": f"Applications file location: {APPLICATIONS_FILE}"
    }


@router.get("/api/download-profile-setups")
async def download_profile_setups():
    """Get path to profile setup Excel file"""
    if not PROFILE_SETUP_FILE.exists():
        raise HTTPException(status_code=404, detail="No profile setups found")
    
    return {
        "success": True,
        "file_path": str(PROFILE_SETUP_FILE),
        "message": f"Profile setup file location: {PROFILE_SETUP_FILE}"
    }
